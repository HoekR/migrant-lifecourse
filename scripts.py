import re
import os
import json
import pandas as pd
from string import Template
import numpy as np
import warnings

warnings.filterwarnings('ignore')

excelfile = pd.ExcelFile("/Users/rikhoekstra/Downloads/Lifecourses NA tentoonstelling.xlsx")

excelfile.sheet_names

names = {'frances_larder':'Frances Larder (NL-AU)',
 'johnny-young':'Johnny Young',
 'adrian_strik':'Adrian Strik (NL-AU)',
 'elly_spillekom':'Elly Spillekom & Dick Koelewijn',
 'karien_dekker':'Karien Dekker (NL-AU)',
 'kurt_van_wijck':'Kurt van Wijck (NL-AU)',
 'fred_feddes':'Fred Feddes (NL-AU_NL)',
 'dianne_biederberg':'Dianne Biederberg (Eastern Euro',
 'helena_meinema':'Helena Meinema (AU-NL)',
 'matthew_meinema':'Matthew Meinema (AU-NL)',
 'jennifer_tucker':'Jennifer Tucker (AU-NL)',
 'joseph_teunissen':'Joseph Teunissen (NL-AU-NL)',
 'petronella_wensing':'Petronella Wensing (NL-AU)',
 'cornelis_bregman':'Cornelis Bregman (NL-AU-NL)',
 'klaas_van_huffelen':'Klaas van Huffelen (NL-AU-NL)',
 'adriana_zevenbergen':'Adriana Zevenbergen (NL-AU)',}

namesfile = './data/names.json'


with open(namesfile,'w') as nm:
    json.dump(names, nm)

# %% [markdown]
# ### Templates

# %%
linktemplate = Template('<a href="$link" target="details">$info</a>')
mcardtmplt = Template('<a href="$card" target="image" rel="noopener"> $ctype</a>')

naa_templ = Template("https://recordsearch.naa.gov.au/SearchNRetrieve/Interface/ViewImage.aspx?B=$nr")

mcard = Template("/$card")

youtubetempl = Template("""<iframe width="420" height="315"
src="https://www.youtube.com/embed/$nr">
</iframe>""")

conversions = {'migration_card':'<img src="/wp-content/uploads/2022/05/mkaart_small.jpg" alt="migrant card thumbnail" />',
               'item_id_naa':'<img src="/wp-content/uploads/2022/05/naa_thmb.jpg" alt="naa recordsearch thumbnail" />',
               'ext_link': '<img src="/wp-content/uploads/2022/05/link_external.png" height="15" width="15" alt="external link" />'
              }

# %%
grid_templ = Template("""<div class="grid-container">
  <div class="personinfo">$personinfo</div>
  <div class="familyinfo">$family</div>
</div>""")

# %%
toptmpt = Template("""<div class="grid-container">
  <div class="personinfo">$personinfo</div>
  <div class="familyinfo">$family</div>
  </div>
  <div class="timeline">$events</div>""")
tmplt = Template("""
 <div class="container $l_or_r">
   <div class="content">
	 <h2>$date</h2>
	 <p>$item</p>
   </div>
 </div>
""")

# %%
container = Template("""<!DOCTYPE html>
<html>
<head/>
<body>
<img class="alignnone wp-image-256" src="/wp-content/uploads/2022/05/puzzle_flags-300x158.png" alt="" width="152" height="80" />
$body
</body>
</html>""")

# %%
def name_in_name(inp, name):
    result = False
    nms = inp.split(' ')
    for n in nms:
        if n.lower() in name.lower():
            result = True
    return result

# %%
import os
def card2local(cardname):
    n = os.path.split(cardname)[-1]
    nn = os.path.splitext(n)[0]
    nnn = nn.split('_')[-2:]
    return '_'.join(nnn)
    

# %%
pat = re.compile("v=(?P<nr>.*)")
def yt2nr(naa, pat=pat):
    nr = naa
    n = pat.search(naa)
    try:
        nr = n.groupdict().get('nr')
    except AttributeError:
        pass
    return nr

# %%
youtubetempl.substitute(nr=yt2nr("https://www.youtube.com/watch?v=4w1mtCmld4s"))

# %%
schemes = {'nama':'NAMA', 
           'ngas':'NGAS', 
           'ex-servicemen': 'empire-and-allied-ex-servicemen-scheme-1948-1955',
           'youth program':'working-holiday-scheme',
           'visa 457': 'working-holiday-scheme',
           'working holiday': 'working-holiday-scheme'}
def concat_evnt(row, cols, links, q, schemes=schemes):
    event = []
    for item in cols:
        e = ''
        if item in row.keys():
            if item == 'event':
                e = f'<div class="event">{row[item]}</div>'
            elif item == 'event_type':
                if row[item].lower() in schemes.keys():
                    e += f'<a href="https://migrant.huygens.knaw.nl/{schemes.get(row[item].lower())}">{row[item]}</a>'
                    #print(e)
            else:
                e = f'{row.get(item)}'
                if pd.isna(e):
                    e=''
            if item in q:
                e = f'<em>{e}</em>'
            elif links.get(item):
                for l in links[item]:
                    try:
                        lnk = row[l]
                        if pd.notna(lnk) and lnk.strip()!='':
                            if l=='source_online':
                                if 'youtube' in lnk:
#                                     print(l, lnk)
                                    ytnr = yt2nr(lnk)
                                    e += youtubetempl.substitute(nr=ytnr)
                                else:
                                    e += linktemplate.substitute(link=lnk, info=f' read more' + ' ' + conversions['ext_link']) #f'<a href="{lnk}">more info</a> <br/>'
                    except KeyError:
                        pass
                for l in links.get('general'):
#                         print(l)
                        try:
                            lnk = row[l]
                            if lnk:
                                if l=='item_id_naa':
                                    lnk = isolate_barcode(lnk)
                                    lnk = naa_templ.substitute(nr=lnk)         
                                elif l=='migration_card':
                                    lnk = card2local(lnk)
                                    lnk = mcard.substitute(card=lnk)
                                e += mcardtmplt.substitute(card=lnk, ctype=conversions.get(item) or 'more info')
#                             if pd.notna(lnk):
#                                 e += linktemplate.substitute(link=lnk, info=f' read more' + ' ' + conversions['ext_link'])
                        except KeyError:
                            pass
        if e != '':
            event.append(e)
    evnt = '</p><p>'.join(event)
    result = f"<p>{evnt}</p>"
    return result 

# %%
import re
pat = re.compile("""(?P<bc>[0-9]{4,10})""")
def isolate_barcode(inp, pat=pat):
    res = pat.search(inp)
    gd = res.groupdict()
    result = gd.get('bc')
    return result
        
    
def concat_famevents(row, cols, links, conversions=conversions):
    event = []
    for item in cols:
        e = ''
        if item in row.keys():
#             print(item)
            txt = row[item]
            if pd.notna(txt):
                if item in links:
                    if txt.strip() != '':
                        if item=='item_id_naa':
                            txt = isolate_barcode(txt)
                            txt = naa_templ.substitute(nr=txt)
                        if item=='migration_card':
                            txt = card2local(txt)
                            txt = mcard.substitute(card=txt)
                        e += mcardtmplt.substitute(card=txt, ctype=conversions.get(item) or 'more info')
                else:
                    e = f' {txt} '
        if e != '':
            event.append(e)
    evnt = '-'.join(event)
    result = f"<li>{evnt}</li>"
    return result

# %%
def mkdate(row):
    date = []
    for part in ['day', 'month', 'year']:
        prt = row.get(part)
        if not pd.isna(prt):
            date.append(int(prt))
    if date:
        try:
            date = [str(d) for d in date]
            result = '-'.join(date)
            return result
        except ValueError:
            print (row, date)

# %%
class Sheet2Timeline(object):
    def __init__(self, name, excelfile):
        strik = excelfile.parse(name)
        self.name = name
        self.df = strik
        self.df = self.df[self.df.event.notna()]
        self.df.event = self.df.event.str.strip()
        if len(self.df)==0:
            print(f'{key} has no records, aborting')
        else:
            self.df.rename(columns=str.lower, inplace=True)
            self.df.rename(columns={c: c.strip().replace(' ', '_') for c in self.df.columns}, inplace=True)
            self.df.rename(columns={c: c.strip().replace('-', '_') for c in self.df.columns}, inplace=True)
            self.df.drop([c for c in self.df.columns if 'unnamed' in c], axis=1, inplace=True)
            #self.df = strik[strik.event_type.notna()]
            for c in ['family_name', 'interpositions', 'first_name']: 
                self.df[c] = self.df[c].fillna('')
            self.df['nm'] = self.df.apply(lambda row: row.family_name + ', ' + row.first_name + ' ' + row.interpositions, axis=1)
            self.eventcols = ['event',
                              'event_type',  
                              'notes',
                              'migrant_quotes', 
                              ]
            self.linkcols = {'notes':['source_online','source'],
                            'general':['migration_card','item_id_naa']}
            self.quotes = ['migrant_quotes']
            for c in ['event_type', 'notes', 'migrant_quotes', 'source_online', 'migration_card','item_id_naa']:
                if c in self.df.columns:
                    self.df[c].fillna('', inplace=True)
  

    def make_event(self):
        self.df['eventtext'] = self.df.apply(lambda row: concat_evnt(row, self.eventcols, self.linkcols, self.quotes) , axis=1)
        return self.df['eventtext']
        
    
    def make_personcard(self):
        profile = ''
        nddf = self.df[self.df.first_name.notna()]
        main_person = nddf[nddf.first_name.apply(lambda x: name_in_name(x, self.name)) & nddf.relation.isna()]
        mainrow = main_person[main_person.event == 'Birthday']
        mainrow.fillna('', inplace=True)
        d = mainrow.to_dict(orient='records')
        prfl = []
        fields = {'nm':'name', 
                      'date':'birthdate',
                      'location':'place of birth', 
                      'nationality':'nationality', 
                      'religion':'religion', 
                      'occupation':'occupation', 
                      'relation':'relation'}
        if len(d)> 0:
            for k in fields:
                e = ''
                val = d[0].get(k)
                if val and pd.notna(val):
                    kval = fields[k]
                    e = f"<dt>{kval.upper()}</dt>"
                    e += f"<dd>{d[0][k]}</dd>"
                prfl.append(e)
            p = '\n\n'.join(prfl)
            profile = f"<dl>{p}</dl>"
        return profile
    
    def make_family(self):
        famevents = self.df[(self.df.event=='Birthday') & (self.df.relation.notna())] 
        fams = famevents.apply(lambda row: concat_famevents(row, ['nm', 'date', 'relation', 'migration_card', 'item_id_naa'],['migration_card', 'item_id_naa']), axis=1)
        if len(fams)==0:
            result = ''
        else:
            result = f'<ul>{"".join(list(fams))}</ul>' #fams.to_string(index=False)
        return result
    
    def sheet2timeline(self):
        for c in ['day', 'month', 'year']:
            self.df[c] = pd.to_numeric(self.df[c], errors='coerce')
        self.df['date'] = self.df.apply(lambda row: mkdate(row), axis=1)
        self.make_event()
        famevents = self.df[(self.df.event=='Birthday') & (self.df.relation.notna())]
        events = self.df[~self.df.index.isin(list(famevents.index))].reset_index()
        return events
#         self.make_familyname()
        
    def render(self, mask=None):
        select = self.sheet2timeline()
        select['lr'] = np.where(select.index % 2==0, 'right', 'left')
        return select
    
    def __repr__(self):
        return self.name

# %%
shts = {}
for key in names:
    ob = Sheet2Timeline(names[key], excelfile)
    if len(ob.df)>0:
            shts[key] = {}
            shts[key]['lifeline']= ob.render()
            shts[key]['profile'] = ob.make_personcard()
            shts[key]['family'] = ob.make_family()



# %%
basedir = '/Users/rikhoekstra/develop/emigratie/tentoonstelling/personen'
for key in names:
    tmplts = []
    frame = shts.get(key)
    if frame and frame.get('lifeline') is not None: 
        if len(frame.get('lifeline'))>0:
            strikdict = frame['lifeline'].to_dict(orient='records')
            for r in strikdict:
                tmplts.append(tmplt.substitute(l_or_r=r.get('lr'),
                                           date=r.get('date'), 
                                           item=r.get('eventtext')))
            events = "\n".join(tmplts)
            prfl = frame['profile']
            family = frame['family']
            body = toptmpt.substitute(events=events, personinfo=prfl, family=family)
            cntnr = container.substitute(body=body)
            with open(os.path.join(basedir, key + '.html'), 'w') as flout:
                flout.write(cntnr)

# %%
tdf = shts['frances_larder']['lifeline']
#tdf[tdf.relation.isna() ]
famevents = tdf[(tdf.event=='Birthday') & (tdf.relation.notna())]
famcols = ['nm', 'date', 'relation', 'migration_card', 'item_id_naa']
famevents.apply(lambda row: concat_famevents(row, famcols, ['migration_card', 'item_id_naa']), axis=1)
tdf[~tdf.index.isin(list(famevents.index))]

# %%
ib = Sheet2Timeline(names['adriana_zevenbergen'], excelfile)
sel = ib.sheet2timeline()
#sel.apply(lambda row: concat_evnt(row, ib.eventcols, ib.linkcols, ib.quotes) , axis=1)
with pd.option_context('display.max_colwidth', None):
  print(sel.apply(lambda row: concat_evnt(row, ib.eventcols, ib.linkcols, ib.quotes) , axis=1))

# %%
with pd.option_context('display.max_colwidth', None):
  print(ib.make_event())


# %%
[c for c in mcards if c!='']

# %%
mcards = []
for key in names:
    ob = Sheet2Timeline(names[key], excelfile)
    if 'migration_card' in ob.df.columns:
        crds = ob.df[ob.df['migration_card'].notna()]['migration_card']
        if len(crds)>0:
            mcards.extend(list(crds))

# %%
ob.df[ob.df['migration_card'].notna()]['migration_card']

# %%
ob = Sheet2Timeline(names['adriana_zevenbergen'], excelfile)

# %%
ob.render()

# %%
ob = Sheet2Timeline(names['elly_spillekom'], excelfile)
ob.df

# %%
from openpyxl import load_workbook
import pandas as pd
from pathlib import Path
src_file = src_file = Path.cwd() / 'Hard Facts.xlsx'

wb = load_workbook(filename = src_file)

# %%
wb.sheetnames

# %%
striksheet = wb['Fred Feddes (NL-AU_NL)']

# %%
striksheet.tables.keys()

# %%
tbl = striksheet.tables['Tabel1']
ref = tbl.ref
ref

# %%
# Access the data in the table range
data = striksheet
rows_list = []

# Loop through each row and get the values in the cells
for row in data:
    # Get a list of all columns in each row
    cols = []
    for col in row:
        cols.append(col.value)
    rows_list.append(cols)

# Create a pandas dataframe from the rows_list.
# The first row is the column names
df = pd.DataFrame(data=rows_list[1:], index=None, columns=rows_list[0])


# %%
import numpy as np
df['lr'] = np.where(df.index % 2 ==0, 'right', 'left')

# %%
df['lr']

# %%
for c in ['day', 'month', 'year']:
    df[c] = pd.to_numeric(df[c], errors='coerce')

# %%
def mkdate(row):
    date = []
    for part in ['day', 'month ', 'year']:
        prt = row.get(part) 
        if not pd.isna(prt):
            date.append(int(prt))
    if date:
        try:
            date = [str(d) for d in date]
        
            result = '-'.join(date)
            return result
        except ValueError:
            print (row, date)

# %%
df['date'] = df.apply(lambda x: mkdate(x), axis=1)

# %%
#keepcols = ['event', 'notes',]

for c in ["migration card", "notes", 'interpositions', 'event']:
    df[c].fillna('', inplace=True)
df['nm'] = df['family name'] + ', ' + df['first name'] + ' ' + df['interpositions']
#df["notes"].fillna('', inplace=True)
df['evnt'] = df['event'] + '<br/>' + df['nm'] 
df['evnt'] = df['evnt']  + '<br/> ' +  df['notes'] + "<br/>" + df['migration card']
df['evnt'] = df['evnt']+ '<br/> '
# def make_dllst(row, keepcols):
#     dllst = ''
#     for item in row[keepcols]:
#         e = row[item]
#         if e:
#             dllst += ' ' + e
#     print (e)
#     result = f'{dllst}'
#     return result

# %%
df['evnt']

# %%
df

# %%
select = df

# %%
select = df[df['first name'].str.contains('Adrian', regex=True)]
select.reset_index()

# %%
import numpy as np
select['lr'] = np.where(select.index % 2 ==0, 'right', 'left')

# %%
df['date'] = df.apply(lambda x: mkdate(x), axis=1)

# %%
strikdict = select.to_dict(orient='records')

# %%
# strikdict = df.to_dict(orient='records')

# %%
toptmpt = Template("""<div class="timeline">$events</div>""")
tmplt = Template("""
 <div class="container $l_or_r">
   <div class="content">
	 <h2>$date</h2>
	 <p>$item</p>
   </div>
 </div>
""")


# %%
strikdict

# %%
tmplts = []

for r in strikdict:
    links = []
#     for item in ['migration card', ]
    tmplts.append(tmplt.substitute(l_or_r=r.get('lr'),
                                   date=r.get('date'), 
                                   item=r.get('evnt')))

# %%
tmplts

# %%
body = toptmpt.substitute(events="\n".join(tmplts))

# %%
container = Template("""<!DOCTYPE html>
<html>
<head><link rel= "stylesheet" type= "text/css" href= "./timeline.css"></head>
<body>
$body
</body>
</html>""").substitute(body=body)

# %%
with open('test.html', 'w') as fn:
    fn.write(container)

# %%
strik = excelfile.parse(names['dianne_biederberg'])
df = strik
df.columns

# %%
df.rename(columns=str.lower, inplace=True)
df.rename(columns={c: c.strip().replace(' ', '_') for c in df.columns}, inplace=True)
df.rename(columns={c: c.strip().replace('-', '_') for c in df.columns}, inplace=True)
df.drop([c for c in df.columns if 'unnamed' in c], axis=1, inplace=True)
for c in ['family_name', 'interpositions', 'first_name']: 
            df[c] = df[c].fillna('')
#df['nm'] = self.df.apply(lambda row: row.family_name + ', ' + row.first_name + ' ' + row.interpositions, axis=1)


# %%



